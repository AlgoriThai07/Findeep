import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

# Load both Excel files
files = [
    ('tesla_10Q.xlsx', 'Tesla'),
    ('jp_morgan_10Q_full.xlsx', 'JP Morgan')
]

filtered_dfs = {}

for file, company in files:
    xls = pd.ExcelFile(file)
    # Assume the summary sheet is the first sheet
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], dtype=object)
    # Filter out columns (tags) and rows (dates) with all nulls
    df = df.dropna(axis=0, how='all')  # Drop rows where all values are null
    df = df.dropna(axis=1, how='all')  # Drop columns where all values are null
    df['Company'] = company
    # Move 'Company' column to the front
    cols = df.columns.tolist()
    if 'Company' in cols:
        cols.insert(0, cols.pop(cols.index('Company')))
        df = df[cols]
    filtered_dfs[company] = df

combined_df = pd.concat(filtered_dfs.values(), ignore_index=True)
# Move 'Company' column to the front in combined_df
cols = combined_df.columns.tolist()
if 'Company' in cols:
    cols.insert(0, cols.pop(cols.index('Company')))
    combined_df = combined_df[cols]

# Write to a new Excel file with separate sheets for each company and a combined sheet
with pd.ExcelWriter('filtered_combined.xlsx', engine='openpyxl') as writer:
    for company, df in filtered_dfs.items():
        df.to_excel(writer, sheet_name=company, index=False)
        worksheet = writer.sheets[company]
        for i, col in enumerate(df.columns, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 25  # Set width to 25
    combined_df.to_excel(writer, sheet_name='Combined', index=False)
    worksheet = writer.sheets['Combined']
    for i, col in enumerate(combined_df.columns, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25  # Set width to 25


# Replace nulls with 'N/A' for export
combined_df_export = combined_df.fillna('N/A')

# Export to CSV
combined_df_export.to_csv('filtered_combined.csv', index=False, encoding='utf-8')


# Export to JSON (flat array of rows, pretty-printed)
combined_df_export.to_json('filtered_combined.json', orient='records', force_ascii=False, indent=2)

print('Filtered and combined data written to filtered_combined.xlsx, filtered_combined.csv, filtered_combined.json')
