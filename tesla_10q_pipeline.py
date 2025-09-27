# --- prerequisites ---
# pip install requests pandas openpyxl python-dateutil

import requests
import pandas as pd
from dateutil import parser as dateparser
from pathlib import Path

CIK = "0001318605"  # Tesla, Inc.
USER_AGENT = "Thai Findeep/1.0 thaiviet0703@gmail.com"
OUTFILE = Path("tesla_10Q.xlsx")

US_GAAP_TAGS = [
    "Revenues",
    "CostOfRevenue",
    "GrossProfit",
    "OperatingExpenses",
    "OperatingIncomeLoss",
    "NetIncomeLoss",
    "EarningsPerShareDiluted",
    "Assets",
    "Liabilities",
    "StockholdersEquity",
    "CashAndCashEquivalentsAtCarryingValue",
    "NetCashProvidedByUsedInOperatingActivities",
    "NetCashProvidedByUsedInInvestingActivities",
    "NetCashProvidedByUsedInFinancingActivities",
]

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT, "Accept-Encoding": "gzip, deflate"})

def get_json(url, timeout=60):
    r = session.get(url, timeout=timeout)
    r.raise_for_status()
    return r.json()

def get_submissions(cik: str) -> dict:
    return get_json(f"https://data.sec.gov/submissions/CIK{cik}.json")

def get_companyfacts(cik: str) -> dict:
    return get_json(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json")

# 1) Latest Tesla 10-Q metadata (filing dates/periods/accessions)

from datetime import datetime, timedelta

subs = get_submissions(CIK)
recent = subs.get("filings", {}).get("recent", {})
forms = recent.get("form", [])
filing_dates = recent.get("filingDate", [])
periods = recent.get("reportDate", [])
accessions = recent.get("accessionNumber", [])

# Calculate the date 10 years ago from today
ten_years_ago = (datetime.now() - timedelta(days=365*10)).date()

ten_q_meta = []
for form, fdate, period, acc in zip(forms, filing_dates, periods, accessions):
    if form == "10-Q":
        # Parse the filing date
        try:
            filing_dt = pd.to_datetime(fdate).date()
        except Exception:
            continue
        if filing_dt >= ten_years_ago:
            ten_q_meta.append({
                "filingDate": fdate,
                "reportPeriod": period or None,
                "accession": acc
            })

ten_q_meta = sorted(ten_q_meta, key=lambda x: x["filingDate"], reverse=True)
allowed_accns = set(x["accession"].replace("-", "") for x in ten_q_meta if x["accession"])

# 2) All Tesla XBRL facts
facts = get_companyfacts(CIK)

def extract_quarterly_values(facts_json, tag):
    """
    Return list of dicts from us-gaap/{tag} for 10-Q entries.
    Keep USD for $ amounts; keep per-share for EPS.
    """
    out = []
    tag_obj = facts_json.get("facts", {}).get("us-gaap", {}).get(tag)
    if not tag_obj:
        return out

    for unit, arr in tag_obj.get("units", {}).items():
        for row in arr:
            form = row.get("form")
            end = row.get("end")
            val = row.get("val")
            accn = row.get("accn")
            if form != "10-Q" or end is None or val is None or accn is None:
                continue

            # units filtering
            if tag == "EarningsPerShareDiluted":
                # EPS comes as per-share; allow any 'per share' unit variant
                if "share" not in unit.lower():
                    continue
            else:
                if unit != "USD":
                    continue

            out.append({
                "tag": tag,
                "form": form,
                "end": end,
                "value": val,
                "uom": unit,
                "accn": accn
            })
    return out

rows = []
for tag in US_GAAP_TAGS:
    rows.extend(extract_quarterly_values(facts, tag))

raw_df = pd.DataFrame(rows)
if raw_df.empty:
    raise SystemExit("No quarterly values found. Check your User-Agent or tags.")

# 3) Keep only entries tied to our latest known 10-Q accessions
raw_df["accn_stripped"] = raw_df["accn"].str.replace("-", "", regex=False)
raw_df = raw_df[raw_df["accn_stripped"].isin(allowed_accns)].copy()

# 4) Build Summary sheet (latest 8 quarter-ends across tags)
raw_df["end_dt"] = pd.to_datetime(raw_df["end"])
latest_quarters = (
    raw_df[["end_dt"]].drop_duplicates().sort_values("end_dt", ascending=False).head(40)["end_dt"].tolist()
)


summary_df = (
    raw_df[raw_df["end_dt"].isin(latest_quarters)]
    .assign(Date=lambda d: d["end_dt"].dt.strftime("%Y-%m-%d"))
    .pivot_table(index="Date", columns="tag", values="value", aggfunc="last")
    .sort_index()
)

summary_df = summary_df[sorted(summary_df.columns)]

# 5) Write Excel

# Delete old Excel file if it exists
if OUTFILE.exists():
    OUTFILE.unlink()


import openpyxl
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(OUTFILE, engine="openpyxl") as xls:
    summary_df.to_excel(xls, sheet_name="Summary")
    worksheet = xls.sheets["Summary"]
    for i, col in enumerate(summary_df.columns, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25
    (raw_df
        .drop(columns=["accn_stripped", "end_dt"])
        .sort_values(["tag","end"])
        .to_excel(xls, sheet_name="RawFacts", index=False))
    worksheet = xls.sheets["RawFacts"]
    for i, col in enumerate(raw_df.drop(columns=["accn_stripped", "end_dt"]).sort_values(["tag","end"]).columns, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25

print(f"Done. Wrote {OUTFILE.resolve()}")
