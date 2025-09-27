import requests
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta

CIK = "0000019617"  # JP Morgan Chase & Co.
USER_AGENT = "Thai Findeep/1.0 thaiviet0703@gmail.com"
OUTFILE = Path("jp_morgan_10Q_full.xlsx")

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT, "Accept-Encoding": "gzip, deflate"})

# Helper to get all 10-Q filings from SEC EDGAR search API

def get_all_10q_accessions(cik, years=15):
    base_url = f"https://www.sec.gov/cgi-bin/browse-edgar"
    params = {
        "action": "getcompany",
        "CIK": cik,
        "type": "10-Q",
        "owner": "exclude",
        "count": 100,
        "output": "atom"
    }
    accessions = []
    cutoff_date = datetime.now() - timedelta(days=365*years)
    start = 0
    while True:
        params["start"] = start
        r = session.get(base_url, params=params)
        r.raise_for_status()
        feed = r.text
        entries = feed.split("<entry>")[1:]
        if not entries:
            break
        for entry in entries:
            # Parse accession number and filing date
            try:
                accn = entry.split("<accession-number>")[1].split("</accession-number>")[0]
                fdate = entry.split("<filing-date>")[1].split("</filing-date>")[0]
                fdate_dt = datetime.strptime(fdate, "%Y-%m-%d")
                if fdate_dt >= cutoff_date:
                    accessions.append({"accession": accn, "filingDate": fdate})
            except Exception:
                continue
        start += 100
    return accessions

# Get all 10-Q accessions for last 15 years
all_10q_meta = get_all_10q_accessions(CIK, years=15)
allowed_accns = set(x["accession"].replace("-", "") for x in all_10q_meta)

# Get XBRL facts (same as before)
def get_companyfacts(cik: str) -> dict:
    return session.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json").json()

US_GAAP_TAGS = [
    "Revenues",
    "CostOfRevenue",
    "GrossProfit",
    "OperatingExpenses",
    "OperatingIncomeLoss",
    "NetIncomeLoss",
    "EarningsPerShareDiluted",
    "Assets",
    "Liabilities",
    "StockholdersEquity",
    "CashAndCashEquivalentsAtCarryingValue",
    "NetCashProvidedByUsedInOperatingActivities",
    "NetCashProvidedByUsedInInvestingActivities",
    "NetCashProvidedByUsedInFinancingActivities",
]

def extract_quarterly_values(facts_json, tag):
    out = []
    tag_obj = facts_json.get("facts", {}).get("us-gaap", {}).get(tag)
    if not tag_obj:
        return out
    for unit, arr in tag_obj.get("units", {}).items():
        for row in arr:
            form = row.get("form")
            end = row.get("end")
            val = row.get("val")
            accn = row.get("accn")
            if form != "10-Q" or end is None or val is None or accn is None:
                continue
            if tag == "EarningsPerShareDiluted":
                if "share" not in unit.lower():
                    continue
            else:
                if unit != "USD":
                    continue
            out.append({
                "tag": tag,
                "form": form,
                "end": end,
                "value": val,
                "uom": unit,
                "accn": accn
            })
    return out

facts = get_companyfacts(CIK)
rows = []
for tag in US_GAAP_TAGS:
    rows.extend(extract_quarterly_values(facts, tag))

raw_df = pd.DataFrame(rows)
if raw_df.empty:
    raise SystemExit("No quarterly values found. Check your User-Agent or tags.")

raw_df["accn_stripped"] = raw_df["accn"].str.replace("-", "", regex=False)
raw_df = raw_df[raw_df["accn_stripped"].isin(allowed_accns)].copy()

raw_df["end_dt"] = pd.to_datetime(raw_df["end"])
all_quarters = raw_df[["end_dt"]].drop_duplicates().sort_values("end_dt", ascending=False)["end_dt"].tolist()


summary_df = (
    raw_df[raw_df["end_dt"].isin(all_quarters)]
    .assign(Date=lambda d: d["end_dt"].dt.strftime("%Y-%m-%d"))
    .pivot_table(index="Date", columns="tag", values="value", aggfunc="last")
    .sort_index()
)

summary_df = summary_df[sorted(summary_df.columns)]

if OUTFILE.exists():
    OUTFILE.unlink()


import openpyxl
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(OUTFILE, engine="openpyxl") as xls:
    summary_df.to_excel(xls, sheet_name="Summary")
    worksheet = xls.sheets["Summary"]
    for i, col in enumerate(summary_df.columns, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25
    (raw_df
        .drop(columns=["accn_stripped", "end_dt"])
        .sort_values(["tag","end"])
        .to_excel(xls, sheet_name="RawFacts", index=False))
    worksheet = xls.sheets["RawFacts"]
    for i, col in enumerate(raw_df.drop(columns=["accn_stripped", "end_dt"]).sort_values(["tag","end"]).columns, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25

print(f"Done. Wrote {OUTFILE.resolve()}")
